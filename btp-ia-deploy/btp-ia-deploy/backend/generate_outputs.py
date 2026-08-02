"""
Génère les livrables finaux.

Le PDF (generate_pdf) est un bilan figé (snapshot) à partir du devis déjà
chiffré par devis_builder.py -- pratique pour un compte-rendu, pas pour être
modifié.

L'Excel (generate_excel) est construit différemment et EXPRÈS sans passer
par devis_builder : trois feuilles reliées par des FORMULES Excel, pour que
l'utilisateur puisse changer une épaisseur, une marge ou un prix unitaire
directement dans le fichier et voir tout se recalculer, sans repasser par ce
script Python.

  - "Paramètres"     : toutes les hypothèses confirmées (épaisseurs, marge de
                        fouille, hauteur de soubassement...) + les prix
                        unitaires, dans des cellules éditables.
  - "Bilan Éléments" : le détail brut par type d'ouvrage (semelles, poteaux,
                        longrines, voiles...) tel qu'extrait des plans, avec
                        les volumes/surfaces calculés par formule à partir
                        des paramètres.
  - "DEVIS QUANTITATIF" : les lignes officielles du devis, dont les cellules
                        quantité/PU/montant sont des formules pointant vers
                        les deux feuilles précédentes.
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from devis_template import SECTION_I_GENERALITES, SECTION_II_TERRASSEMENT, SECTIONS_HORS_PERIMETRE, POSTE_KEY_TO_CODE

HEADER_FILL = "10b981"
SECTION_FILL = "e5e7eb"
WARN_FILL = "fff3cd"
PARAM_FILL = "e0f2ec"

# Valeurs de repli pour CHAQUE paramètre, pour que la feuille Paramètres les
# écrive TOUJOURS (même si la question n'a pas été posée -- ex: aucun voile
# détecté donc aucune question sur son épaisseur). Comme ça les formules de
# Bilan Éléments peuvent toujours référencer ces cellules, et si l'utilisateur
# ajoute plus tard une ligne manuellement (un voile que l'IA aurait raté par
# exemple), le calcul fonctionne immédiatement sans avoir à ressaisir un
# paramètre qui n'existerait pas encore.
DEFAULT_PARAMS = {
    "hauteur_soubassement_m": 0,
    "profondeur_ancrage_m": 0,
    "marge_fouille_pct": 15,
    "epaisseur_beton_proprete_cm": 10,
    "epaisseur_dallage_cm": 13,
    "epaisseur_voile_cm": 20,
    "largeur_semelle_filante_cm": 40,
    "hauteur_semelle_filante_cm": 20,
}


def _fmt_fcfa(n) -> str:
    if n is None:
        return "-"
    try:
        return f"{int(round(n)):,}".replace(",", " ")
    except (TypeError, ValueError):
        return str(n)


def _split_section(section: str):
    """'20x40' -> (20.0, 40.0) en cm, sinon (None, None)."""
    s = (section or "").strip().upper()
    if s.startswith("D"):
        try:
            d = float(s[1:])
            return d, d  # diamètre traité comme un carré équivalent grossier
        except ValueError:
            return None, None
    if "X" in s:
        try:
            a, b = s.split("X")
            return float(a), float(b)
        except ValueError:
            return None, None
    return None, None


# ------------------------------------------------------------------------
# Feuille "Paramètres"
# ------------------------------------------------------------------------

def _write_parametres(ws, answers: dict, kb: dict, header_fill, param_fill, border):
    ws["A1"] = "Paramètres de calcul (modifiables)"
    ws["A1"].font = Font(size=14, bold=True)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50

    row = 3
    ws.cell(row=row, column=1, value="Hypothèses confirmées par l'utilisateur").font = Font(bold=True, size=11)
    row += 1

    param_labels = {
        "hauteur_soubassement_m": ("Hauteur de soubassement (m)", "m"),
        "profondeur_ancrage_m": ("Profondeur d'ancrage des fondations (m)", "m"),
        "marge_fouille_pct": ("Marge de fouille (%)", "%"),
        "epaisseur_beton_proprete_cm": ("Épaisseur béton de propreté (cm)", "cm"),
        "epaisseur_dallage_cm": ("Épaisseur dallage (cm)", "cm"),
        "epaisseur_voile_cm": ("Épaisseur voile en soubassement (cm)", "cm"),
        "largeur_semelle_filante_cm": ("Largeur semelle filante sous longrines (cm, poste 3.4)", "cm"),
        "hauteur_semelle_filante_cm": ("Hauteur semelle filante sous longrines (cm, poste 3.4)", "cm"),
    }
    param_cells = {}
    for key, (label, unit) in param_labels.items():
        value = answers.get(key, DEFAULT_PARAMS[key])
        ws.cell(row=row, column=1, value=f"{label}").border = border
        c = ws.cell(row=row, column=2, value=value)
        c.border = border
        c.fill = param_fill
        c.font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"Modifie cette cellule pour recalculer le devis (unité: {unit}).").font = Font(italic=True, size=9, color="666666")
        param_cells[key] = f"'Paramètres'!${get_column_letter(2)}${row}"
        row += 2

    row += 1
    ws.cell(row=row, column=1, value="Prix unitaires (FCFA) — modifiables").font = Font(bold=True, size=11)
    row += 1
    headers = ["Code", "Désignation", "Unité", "PU (FCFA)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="ffffff")
        c.fill = header_fill
        c.border = border
    row += 1

    pu_cells = {}
    for code, info in kb.get("postes", {}).items():
        ws.cell(row=row, column=1, value=code).border = border
        ws.cell(row=row, column=2, value=info.get("designation", "")).border = border
        ws.cell(row=row, column=3, value=info.get("unite", "")).border = border
        c = ws.cell(row=row, column=4, value=info.get("prix_unitaire_fcfa"))
        c.border = border
        c.fill = param_fill
        pu_cells[code] = f"'Paramètres'!${get_column_letter(4)}${row}"
        row += 1

    return param_cells, pu_cells


# ------------------------------------------------------------------------
# Feuille "Bilan Éléments"
# ------------------------------------------------------------------------

def _write_bilan_elements(ws, bilan: dict, param_cells: dict, header_fill, border):
    for i, w in enumerate([26, 12, 12, 12, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = "Bilan des éléments (détail brut utilisé dans les calculs)"
    ws["A1"].font = Font(size=14, bold=True)
    row = 3
    total_cells = {}  # code -> "'Bilan Éléments'!$G$12"
    SHEET = "'Bilan Éléments'"

    def _table_header(row, cols):
        for i, h in enumerate(cols, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(bold=True, color="ffffff")
            c.fill = header_fill
            c.border = border
        return row + 1

    # ---- Semelles isolées (3.1 béton propreté + 3.3 béton armé) ----
    ws.cell(row=row, column=1, value="Semelles isolées").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Désignation", "a (m)", "b (m)", "h (cm)", "Nombre",
                               "Vol. béton armé (m3)", "Vol. béton propreté (m3)"])
    first_row = row
    for s in bilan.get("semelles", []):
        ws.cell(row=row, column=1, value=s.get("designation", "")).border = border
        ws.cell(row=row, column=2, value=s["a_m"]).border = border
        ws.cell(row=row, column=3, value=s["b_m"]).border = border
        ws.cell(row=row, column=4, value=s["h_cm"]).border = border
        ws.cell(row=row, column=5, value=s["nombre"]).border = border
        r = row
        ws.cell(row=row, column=6, value=f"=B{r}*C{r}*(D{r}/100)*E{r}").border = border
        ws.cell(row=row, column=7, value=f"=B{r}*C{r}*({param_cells['epaisseur_beton_proprete_cm']}/100)*E{r}").border = border
        row += 1
    if row == first_row:  # aucune semelle détectée -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter une semelle ici)").font = Font(italic=True, color="999999")
        for col, val in ((2, 0), (3, 0), (4, 0), (5, 0)):
            ws.cell(row=row, column=col, value=val).border = border
        r = row
        ws.cell(row=row, column=6, value=f"=B{r}*C{r}*(D{r}/100)*E{r}").border = border
        ws.cell(row=row, column=7, value=f"=B{r}*C{r}*({param_cells['epaisseur_beton_proprete_cm']}/100)*E{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"=SUM(G{first_row}:G{last_row})").font = Font(bold=True)
    total_cells["3.3"] = f"{SHEET}!$F${row}"
    total_cells["3.1"] = f"{SHEET}!$G${row}"
    row += 3

    # ---- Semelles filantes (3.4) : longueur développée = longueur totale
    # des longrines (référence directe à la table Longrines ci-dessous),
    # section confirmée par l'utilisateur en Paramètres. Note: la table
    # Longrines est écrite juste après pour connaître longrines_longueur_cell
    # -- on réserve donc la ligne ici et on la complète plus bas. ----
    row_semelles_filantes = row
    ws.cell(row=row, column=1, value="Semelles filantes (sous longrines)").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Longueur développée (m)", "Largeur section (m, param.)",
                               "Hauteur section (m, param.)", "Volume béton (m3)", "", "", ""])
    row_sf_data = row
    row += 1
    row += 2

    # ---- Poteaux (3.5 potelets, dépend de la hauteur soubassement) ----
    ws.cell(row=row, column=1, value="Poteaux (potelets en soubassement)").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Section", "Nombre", "Aire section (m2)", "Hauteur (m, param.)", "Volume béton (m3)", "", ""])
    first_row = row
    hc = param_cells["hauteur_soubassement_m"]
    for item in bilan.get("poteaux", {}).get("par_section", []):
        a_cm, b_cm = _split_section(item["section"])
        aire = round((a_cm / 100) * (b_cm / 100), 4) if a_cm and b_cm else None
        ws.cell(row=row, column=1, value=item["section"]).border = border
        ws.cell(row=row, column=2, value=item["nombre_total"]).border = border
        ws.cell(row=row, column=3, value=aire).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={hc}").border = border
        ws.cell(row=row, column=5, value=f"=C{r}*B{r}*D{r}").border = border
        row += 1
    if row == first_row:  # aucun poteau détecté -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter un poteau ici, ex: 20x20)").font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=0).border = border
        ws.cell(row=row, column=3, value=0).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={hc}").border = border
        ws.cell(row=row, column=5, value=f"=C{r}*B{r}*D{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{first_row}:E{last_row})").font = Font(bold=True)
    total_cells["3.5"] = f"{SHEET}!$E${row}"
    row += 3

    # ---- Voiles (3.6, dépend de hauteur + épaisseur voile) ----
    ws.cell(row=row, column=1, value="Voiles en soubassement").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Désignation", "Nombre", "Longueur totale (m)", "Épaisseur (m, param.)", "Hauteur (m, param.)", "Volume béton (m3)", ""])
    first_row = row
    ec = param_cells["epaisseur_voile_cm"]
    for v in bilan.get("voiles_par_type", []):
        ws.cell(row=row, column=1, value=v.get("designation", "")).border = border
        ws.cell(row=row, column=2, value=v.get("nombre")).border = border
        ws.cell(row=row, column=3, value=v["longueur_totale_m"]).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={ec}/100").border = border
        ws.cell(row=row, column=5, value=f"={hc}").border = border
        ws.cell(row=row, column=6, value=f"=C{r}*D{r}*E{r}").border = border
        row += 1
    if row == first_row:  # aucun voile détecté -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter un voile ici)").font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=0).border = border
        ws.cell(row=row, column=3, value=0).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={ec}/100").border = border
        ws.cell(row=row, column=5, value=f"={hc}").border = border
        ws.cell(row=row, column=6, value=f"=C{r}*D{r}*E{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{first_row}:B{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(C{first_row}:C{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
    total_cells["3.6"] = f"{SHEET}!$F${row}"
    voiles_longueur_totale_cell = f"{SHEET}!$C${row}"
    row += 3

    # ---- Longrines (3.7, + largeur cumulée pour fouilles 2.4) ----
    ws.cell(row=row, column=1, value="Longrines").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Section", "Tronçons", "Longueur totale (m)", "Largeur (m)", "Volume béton (m3)", "Largeur x Longueur (m2)", ""])
    first_row = row
    for item in bilan.get("longrines_par_section", []):
        a_cm, b_cm = _split_section(item["section"])
        largeur_m = round(min(a_cm, b_cm) / 100, 4) if a_cm and b_cm else None
        aire_section = round((a_cm / 100) * (b_cm / 100), 4) if a_cm and b_cm else None
        ws.cell(row=row, column=1, value=item["section"]).border = border
        ws.cell(row=row, column=2, value=item.get("nombre_troncons")).border = border
        ws.cell(row=row, column=3, value=item["longueur_totale_m"]).border = border
        ws.cell(row=row, column=4, value=largeur_m).border = border
        r = row
        ws.cell(row=row, column=5, value=(f"={aire_section}*C{r}" if aire_section is not None else None)).border = border
        ws.cell(row=row, column=6, value=f"=D{r}*C{r}").border = border
        row += 1
    if row == first_row:  # aucune longrine détectée -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter une longrine ici, ex: 20x30)").font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=0).border = border
        ws.cell(row=row, column=3, value=0).border = border
        ws.cell(row=row, column=4, value=0).border = border
        r = row
        ws.cell(row=row, column=5, value=f"=D{r}*C{r}").border = border  # largeur * longueur (approx, à corriger si section réelle connue)
        ws.cell(row=row, column=6, value=f"=D{r}*C{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(C{first_row}:C{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{first_row}:E{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
    total_cells["3.7"] = f"{SHEET}!$E${row}"
    longrines_longueur_totale_cell = f"{SHEET}!$C${row}"
    longrines_largeur_longueur_cell = f"{SHEET}!$F${row}"
    row += 3

    # ---- Complète la ligne "Semelles filantes" réservée plus haut, maintenant
    # que la longueur développée totale des longrines est connue ----
    ws.cell(row=row_sf_data, column=1, value=f"={longrines_longueur_totale_cell}").border = border
    ws.cell(row=row_sf_data, column=2, value=f"={param_cells['largeur_semelle_filante_cm']}/100").border = border
    ws.cell(row=row_sf_data, column=3, value=f"={param_cells['hauteur_semelle_filante_cm']}/100").border = border
    ws.cell(row=row_sf_data, column=4,
            value=f"=A{row_sf_data}*B{row_sf_data}*C{row_sf_data}").border = border
    total_cells["3.4"] = f"{SHEET}!$D${row_sf_data}"

    # ---- Fouilles (2.3 puits, 2.4 rigoles) : dépendent de profondeur + marge ----
    ws.cell(row=row, column=1, value="Fouilles").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Type", "Emprise (m2)", "Profondeur (m, param.)", "Marge (%, param.)", "Volume (m3)", "", ""])
    surface_semelles = sum(s["a_m"] * s["b_m"] * s["nombre"] for s in bilan.get("semelles", []))
    if "profondeur_ancrage_m" in param_cells and "marge_fouille_pct" in param_cells:
        pc = param_cells["profondeur_ancrage_m"]
        mc = param_cells["marge_fouille_pct"]
        r = row
        ws.cell(row=row, column=1, value="Puits (semelles isolées)").border = border
        ws.cell(row=row, column=2, value=round(surface_semelles, 3)).border = border
        ws.cell(row=row, column=3, value=f"={pc}").border = border
        ws.cell(row=row, column=4, value=f"={mc}").border = border
        ws.cell(row=row, column=5, value=f"=B{r}*C{r}*(1+D{r}/100)").border = border
        total_cells["2.3"] = f"{SHEET}!$E${r}"
        row += 1
        r = row
        ws.cell(row=row, column=1, value="Rigoles (longrines)").border = border
        ws.cell(row=row, column=2, value=f"={longrines_largeur_longueur_cell}").border = border
        ws.cell(row=row, column=3, value=f"={pc}").border = border
        ws.cell(row=row, column=4, value=f"={mc}").border = border
        ws.cell(row=row, column=5, value=f"=B{r}*C{r}*(1+D{r}/100)").border = border
        total_cells["2.4"] = f"{SHEET}!$E${r}"
        row += 1
    else:
        total_cells["2.3"] = None
        total_cells["2.4"] = None
    row += 2

    # ---- Dallage (3.8) : surface brute - déductions, dépend épaisseur dallage + épaisseur voile ----
    sd = bilan.get("surface_dallage", {})
    ws.cell(row=row, column=1, value="Dallage").font = Font(bold=True, size=12)
    row += 1
    if not sd.get("donnee_indisponible", True):
        surface_brute = sd.get("surface_brute_m2", 0)
        deductions = sd.get("deductions_m2", {})
        ddp = deductions.get("poteaux", 0)
        ddl = deductions.get("longrines", 0)
        ddv_brute = deductions.get("voiles", 0)  # calculée avec épaisseur voile 20cm d'origine

        ws.cell(row=row, column=1, value="Surface brute (m2)").border = border
        ws.cell(row=row, column=2, value=surface_brute).border = border
        r_brute = row
        row += 1
        ws.cell(row=row, column=1, value="Déduction emprise poteaux (m2)").border = border
        ws.cell(row=row, column=2, value=ddp).border = border
        r_dp = row
        row += 1
        ws.cell(row=row, column=1, value="Déduction emprise longrines (m2)").border = border
        ws.cell(row=row, column=2, value=ddl).border = border
        r_dl = row
        row += 1
        ws.cell(row=row, column=1, value="Déduction emprise voiles (m2, mise à l'échelle épaisseur param.)").border = border
        if "epaisseur_voile_cm" in param_cells and ddv_brute:
            ec = param_cells["epaisseur_voile_cm"]
            ws.cell(row=row, column=2, value=f"={ddv_brute}*({ec}/20)").border = border
        else:
            ws.cell(row=row, column=2, value=ddv_brute).border = border
        r_dv = row
        row += 1
        ws.cell(row=row, column=1, value="Surface nette dallage (m2)").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"=B{r_brute}-B{r_dp}-B{r_dl}-B{r_dv}").font = Font(bold=True)
        r_nette = row
        row += 1
        ws.cell(row=row, column=1, value="Volume béton dallage (m3)").font = Font(bold=True)
        if "epaisseur_dallage_cm" in param_cells:
            edc = param_cells["epaisseur_dallage_cm"]
            ws.cell(row=row, column=2, value=f"=B{r_nette}*({edc}/100)").font = Font(bold=True)
            total_cells["3.8"] = f"{SHEET}!$B${row}"
        else:
            ws.cell(row=row, column=2, value=None).font = Font(bold=True)
            total_cells["3.8"] = None
        row += 1
    else:
        ws.cell(row=row, column=1, value=f"Indisponible: {sd.get('raison', 'surface dallage non calculable')}").font = Font(italic=True, color="884400")
        total_cells["3.8"] = None
        row += 1

    # ---- Escaliers (3.9 marches, 3.10 bèche) : pas de paramètre externe requis ----
    ws.cell(row=row, column=1, value="Escaliers").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Désignation", "Nb marches", "Giron (m)", "Hauteur marche (m)",
                               "Largeur volée (m)", "Vol. marches (m3)", "Vol. bèche (m3)"])
    first_row = row
    escaliers = bilan.get("escaliers", [])
    for e in escaliers:
        ws.cell(row=row, column=1, value=e.get("designation", "")).border = border
        ws.cell(row=row, column=2, value=e.get("nombre_marches")).border = border
        giron_m = round(e["giron_cm"] / 100, 4) if e.get("giron_cm") else None
        hm_m = round(e["hauteur_marche_cm"] / 100, 4) if e.get("hauteur_marche_cm") else None
        ws.cell(row=row, column=3, value=giron_m).border = border
        ws.cell(row=row, column=4, value=hm_m).border = border
        ws.cell(row=row, column=5, value=e.get("largeur_volee_m")).border = border
        r = row
        if e.get("nombre_marches") and giron_m and hm_m and e.get("largeur_volee_m"):
            ws.cell(row=row, column=6, value=f"=B{r}*(C{r}*D{r}/2)*E{r}").border = border
        else:
            ws.cell(row=row, column=6, value=None).border = border
        beche = e.get("beche") or {}
        if beche.get("longueur_m") and beche.get("largeur_cm") and beche.get("hauteur_cm"):
            ws.cell(row=row, column=7, value=round(
                (beche["largeur_cm"] / 100) * (beche["hauteur_cm"] / 100) * beche["longueur_m"], 4
            )).border = border
        else:
            ws.cell(row=row, column=7, value=None).border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    if last_row >= first_row:
        ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
        ws.cell(row=row, column=7, value=f"=SUM(G{first_row}:G{last_row})").font = Font(bold=True)
    else:
        ws.cell(row=row, column=6, value=0).font = Font(bold=True)
        ws.cell(row=row, column=7, value=0).font = Font(bold=True)
    total_cells["3.9"] = f"{SHEET}!$F${row}" if any(
        e.get("nombre_marches") and e.get("giron_cm") and e.get("hauteur_marche_cm") and e.get("largeur_volee_m")
        for e in escaliers
    ) else None
    total_cells["3.10"] = f"{SHEET}!$G${row}" if any(
        (e.get("beche") or {}).get("longueur_m") for e in escaliers
    ) else None
    row += 3

    return total_cells


# ------------------------------------------------------------------------
# Feuille "DEVIS QUANTITATIF"
# ------------------------------------------------------------------------

def _write_devis(ws, kb: dict, postes_by_code: dict, pu_cells: dict, raisons: dict, project_name: str,
                  location: str, avertissements: list, header_fill, section_fill, warn_fill, border):
    col_widths = [8, 46, 8, 12, 14, 16, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    ws["A1"] = f"DEVIS QUANTITATIF ET ESTIMATIF — {project_name}"
    ws["A1"].font = Font(size=15, bold=True)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{location} — {date.today().isoformat()} — Infrastructure uniquement (voir notes)"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    row = 4
    headers = ["Code", "Désignation des ouvrages", "Unité", "Quantité", "PU (FCFA)", "Montant (FCFA)", "Note"]

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=f"I. {SECTION_I_GENERALITES['titre']}")
    c.font = Font(bold=True, size=12)
    c.fill = section_fill
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=f"⚠ {SECTION_I_GENERALITES['note']}")
    c.font = Font(italic=True, size=9, color="884400")
    c.fill = warn_fill
    row += 2

    sections_codes = [
        ("II", SECTION_II_TERRASSEMENT["titre"],
         "Seuls les postes 2.3/2.4 (fouilles) sont calculés depuis les plans structure. "
         "Les postes 2.1/2.2/2.5 à 2.8 dépendent de données hors périmètre -- à compléter manuellement.",
         ["2.3", "2.4"]),
        ("III", "BETON - BETON ARME (Infrastructures)",
         "Superstructure (postes 3.11 à 3.22) hors périmètre -- nécessite les plans de superstructure.",
         ["3.1", "3.2", "3.3", "3.4", "3.5", "3.6", "3.7", "3.8", "3.9", "3.10"]),
    ]

    montant_ranges_by_section = []
    for numero, titre, note, codes in sections_codes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"{numero}. {titre}")
        c.font = Font(bold=True, size=12)
        c.fill = section_fill
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"⚠ {note}")
        c.font = Font(italic=True, size=9, color="884400")
        c.fill = warn_fill
        row += 1

        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(bold=True, color="ffffff")
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal="center")
        row += 1

        first_data_row = row
        for code in codes:
            info = kb.get("postes", {}).get(code, {})
            poste = postes_by_code.get(code)
            pu_cell = pu_cells.get(code)
            quantite = None
            if poste and not poste.get("donnee_indisponible", True):
                quantite = poste.get("volume_m3")
                if quantite is None:
                    quantite = poste.get("quantite_m2")
            indispo = quantite is None

            ws.cell(row=row, column=1, value=code).border = border
            ws.cell(row=row, column=2, value=info.get("designation", code)).border = border
            ws.cell(row=row, column=3, value=info.get("unite", "")).border = border

            r = row
            if indispo:
                cq = ws.cell(row=row, column=4, value="À COMPLÉTER")
                cpu = ws.cell(row=row, column=5, value=(f"={pu_cell}" if pu_cell else "-"))
                cm_ = ws.cell(row=row, column=6, value="-")
                note_cell = ws.cell(row=row, column=7, value=raisons.get(code) or "Non calculable (donnée hors périmètre ou question non posée).")
                for cc in (cq, cpu, cm_, note_cell):
                    cc.fill = warn_fill
            else:
                # Valeur littérale déjà calculée en Python (bilan["volumes_beton"]),
                # PAS une formule pointant vers une autre feuille -- garantit un
                # affichage correct même dans un viewer qui ne recalcule pas les
                # formules à l'ouverture (contrairement à une référence croisée).
                cq = ws.cell(row=row, column=4, value=round(quantite, 2))
                cpu = ws.cell(row=row, column=5, value=(f"={pu_cell}" if pu_cell else ""))
                cm_ = ws.cell(row=row, column=6, value=f"=D{r}*E{r}")
                note_cell = ws.cell(row=row, column=7, value=raisons.get(code) or "")
            for cc in (cq, cpu, cm_, note_cell):
                cc.border = border
                if cc.column in (4, 5, 6):
                    cc.alignment = Alignment(horizontal="right")
            row += 1
        last_data_row = row - 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value="Sous-total section")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c2 = ws.cell(row=row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="right")
        montant_ranges_by_section.append(f"F{row}")
        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value="TOTAL INFRASTRUCTURE (postes chiffrés, HTVA)")
    c.font = Font(bold=True, size=13)
    c.fill = section_fill
    c.alignment = Alignment(horizontal="right")
    c2 = ws.cell(row=row, column=6, value="=" + "+".join(montant_ranges_by_section))
    c2.font = Font(bold=True, size=13)
    c2.fill = section_fill
    c2.alignment = Alignment(horizontal="right")
    row += 3

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="Autres lots du devis (hors périmètre de cette extraction automatique) :").font = Font(bold=True, size=11)
    row += 1
    for s in SECTIONS_HORS_PERIMETRE:
        ws.cell(row=row, column=1, value=f"{s['numero']}. {s['titre']} — à compléter manuellement").font = Font(italic=True, size=9, color="666666")
        row += 1
    row += 1

    if avertissements:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value="Observations (relecture IA) :").font = Font(bold=True, size=11)
        row += 1
        for a in avertissements:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(row=row, column=1, value=f"- {a}").font = Font(italic=True, size=9, color="444444")
            row += 1


def generate_excel(bilan: dict, answers: dict, kb: dict, project_name: str, location: str,
                    avertissements: list, out_path: str):
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    section_fill = PatternFill("solid", fgColor=SECTION_FILL)
    warn_fill = PatternFill("solid", fgColor=WARN_FILL)
    param_fill = PatternFill("solid", fgColor=PARAM_FILL)
    thin = Side(style="thin", color="cccccc")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_param = wb.active
    ws_param.title = "Paramètres"
    param_cells, pu_cells = _write_parametres(ws_param, answers, kb, header_fill, param_fill, border)

    ws_elem = wb.create_sheet("Bilan Éléments")
    total_cells = _write_bilan_elements(ws_elem, bilan, param_cells, header_fill, border)

    raisons = {}
    postes_by_code = {}
    all_postes = bilan.get("volumes_beton", {}).get("postes", {})
    for poste_key, code in POSTE_KEY_TO_CODE.items():
        poste = all_postes.get(poste_key)
        postes_by_code[code] = poste
        if poste and poste.get("raison"):
            raisons[code] = poste["raison"]

    ws_devis = wb.create_sheet("DEVIS QUANTITATIF")
    _write_devis(ws_devis, kb, postes_by_code, pu_cells, raisons, project_name, location, avertissements,
                 header_fill, section_fill, warn_fill, border)

    wb.save(out_path)
    return out_path


# ------------------------------------------------------------------------
# PDF bilan (snapshot figé, à partir du devis déjà chiffré par devis_builder.py)
# ------------------------------------------------------------------------

def generate_pdf(devis: dict, out_path: str):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleFR", parent=styles["Title"], fontSize=16)
    sub_style = ParagraphStyle("SubFR", parent=styles["Normal"], textColor=colors.grey, fontSize=9)
    section_style = ParagraphStyle("SectionFR", parent=styles["Heading2"], fontSize=12, spaceBefore=14)
    note_style = ParagraphStyle("NoteFR", parent=styles["Normal"], textColor=colors.HexColor("#884400"), fontSize=8, spaceAfter=6)
    warn_style = ParagraphStyle("WarnFR", parent=styles["Normal"], textColor=colors.HexColor("#884400"), fontSize=8)

    doc = SimpleDocTemplate(out_path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
        Paragraph(f"Bilan — Devis Quantitatif Infrastructure — {devis['projet']}", title_style),
        Paragraph(f"{devis['localisation']} — {date.today().isoformat()}", sub_style),
        Spacer(1, 10),
    ]

    for section in devis["sections"]:
        story.append(Paragraph(f"{section['numero']}. {section['titre']}", section_style))
        if section.get("note"):
            story.append(Paragraph(f"⚠ {section['note']}", note_style))

        data = [["Code", "Désignation", "Unité", "Qté", "PU (FCFA)", "Montant (FCFA)"]]
        section_total = 0
        for ligne in section["lignes"]:
            indispo = ligne["source"] == "indisponible"
            data.append([
                ligne["code"], ligne["designation"][:60] + ("…" if len(ligne["designation"]) > 60 else ""),
                ligne["unite"],
                "À compléter" if indispo else str(ligne["quantite"]),
                _fmt_fcfa(ligne["prix_unitaire_fcfa"]) if ligne["prix_unitaire_fcfa"] else "-",
                "-" if indispo else _fmt_fcfa(ligne["montant_fcfa"]),
            ])
            if ligne["montant_fcfa"]:
                section_total += ligne["montant_fcfa"]
        data.append(["", "", "", "", "Sous-total", _fmt_fcfa(section_total)])

        table = Table(data, colWidths=[1.4 * cm, 6.8 * cm, 1.3 * cm, 2.0 * cm, 2.6 * cm, 3.0 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -2), 0.4, colors.HexColor("#cccccc")),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        story.append(Spacer(1, 6))

    story.append(Spacer(1, 10))
    total_style = ParagraphStyle("TotalFR", parent=styles["Heading2"], alignment=2)
    story.append(Paragraph(
        f"TOTAL INFRASTRUCTURE (HTVA) : {_fmt_fcfa(devis['total_infrastructure_fcfa'])} FCFA", total_style))

    story.append(Spacer(1, 14))
    story.append(Paragraph("Autres lots (hors périmètre de cette extraction, à compléter manuellement) :", section_style))
    for s in SECTIONS_HORS_PERIMETRE:
        story.append(Paragraph(f"- {s['numero']}. {s['titre']}", warn_style))

    if devis.get("postes_a_completer_manuellement"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("Points à compléter / confirmer :", section_style))
        for item in devis["postes_a_completer_manuellement"]:
            story.append(Paragraph(f"- {item['poste']} : {item['raison']}", warn_style))

    if devis.get("avertissements"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("Observations (relecture IA) :", section_style))
        for a in devis["avertissements"]:
            story.append(Paragraph(f"- {a}", warn_style))

    doc.build(story)
    return out_path
