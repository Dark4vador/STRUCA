import uuid
import shutil
import subprocess
import threading
import json
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

from pipeline import run_pipeline
from missing_info import (
    detect_missing_questions, resolve_answer, apply_answers_to_bilan,
    extract_value_from_attachment, _normalise_type, EXCLUS_RESEAU_LONGRINES,
)
from gemini_client import GeminiError
from devis_builder import build_devis
from groq_client import review_devis
from generate_outputs import generate_excel, generate_pdf

BASE_DIR = Path(__file__).parent
UPLOADS_DIR = BASE_DIR / "uploads"
OUTPUTS_DIR = BASE_DIR / "outputs"
UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

KNOWLEDGE_BASE = json.loads((BASE_DIR / "knowledge_base.json").read_text(encoding="utf-8"))


def _recalculate_xlsx_with_libreoffice(xlsx_path: Path) -> bool:
    """Force le recalcul + la mise en cache des formules du classeur en le
    faisant passer par LibreOffice headless (convert-to xlsx sur lui-même).
    Renvoie True si le recalcul a réussi, False sinon (voir v38 plus bas:
    le résultat est utilisé pour rendre un échec visible DANS le fichier
    lui-même, pas seulement dans les logs serveur que l'utilisateur ne voit
    jamais).

    v27 -- openpyxl écrit les FORMULES ('=Bilan Éléments'!$E$55...) mais ne
    les calcule jamais lui-même: sans cette étape, les cellules Quantité/PU/
    Montant de 'DEVIS QUANTITATIF' n'ont AUCUNE valeur mise en cache tant que
    le fichier n'a pas été ouvert au moins une fois dans un vrai tableur.
    Résultat pour l'utilisateur: des cellules qui semblent vides dans
    n'importe quel visualiseur qui ne recalcule pas lui-même (aperçu rapide,
    certains lecteurs web...), même si le fichier n'a rien de cassé -- les
    formules elles-mêmes restent correctes et éditables normalement, cette
    étape ne fait qu'ajouter la valeur mise en cache par-dessus.

    Si LibreOffice n'est pas installé sur cet environnement de déploiement,
    on continue sans bloquer la génération -- le fichier reste utilisable
    dans Excel/LibreOffice (qui recalculent à l'ouverture), seul l'aperçu
    dans un visualiseur non-calculant restera vide."""
    try:
        import tempfile
        with tempfile.TemporaryDirectory() as tmp_dir:
            result = subprocess.run(
                ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", tmp_dir, str(xlsx_path)],
                capture_output=True, text=True, timeout=90,
            )
            if result.returncode != 0:
                print(f"[recalc xlsx] LibreOffice a échoué (code {result.returncode}): "
                      f"{result.stderr[:500]} -- fichier conservé tel quel (formules non mises en cache).")
                return False
            recalculated = Path(tmp_dir) / xlsx_path.name
            if recalculated.exists():
                shutil.move(str(recalculated), str(xlsx_path))
                return True
            else:
                print(f"[recalc xlsx] Fichier recalculé introuvable ({recalculated}) -- fichier conservé tel quel.")
                return False
    except FileNotFoundError:
        print("[recalc xlsx] LibreOffice (soffice) introuvable sur cet environnement -- "
              "fichier conservé tel quel (formules non mises en cache, mais correctes: "
              "Excel/LibreOffice les recalculeront normalement à l'ouverture).")
        return False
    except subprocess.TimeoutExpired:
        print("[recalc xlsx] Délai dépassé lors du recalcul LibreOffice -- fichier conservé tel quel.")
        return False


def _insert_recalc_warning_banner(xlsx_path: Path) -> None:
    """v38 -- si le recalcul LibreOffice a échoué (souvent: LibreOffice pas
    installé sur l'environnement de déploiement), les cellules Quantité/PU/
    Montant du fichier livré à l'utilisateur restent vides tant qu'il ne
    l'ouvre pas lui-même dans un vrai tableur -- un échec qui ne se voyait
    jusqu'ici que dans les logs serveur (jamais consultés par l'utilisateur
    final). On rend l'échec visible EN CLAIR, en tête de la feuille
    'DEVIS QUANTITATIF' elle-même, pour qu'il soit impossible à manquer.

    v47 -- bug corrigé: ws.insert_rows(3) décalait bien les VALEURS des
    lignes vers le bas, mais PAS les zones fusionnées (merge_cells) créées
    par _write_devis pour les titres de section -- une limitation connue
    d'openpyxl. Résultat: après l'insertion, certaines zones fusionnées
    d'origine se retrouvaient décalées d'une ligne par rapport aux vraies
    données, et TOUTE cellule qui tombait alors dans une zone fusionnée
    (autre que la cellule en haut à gauche) était automatiquement vidée par
    Excel/openpyxl -- exactement le symptôme observé (designation/unité/
    quantité/PU vides sur des lignes isolées comme 2.4 ou 3.15, alors que
    le montant, lui, restait correct car pas dans le chemin de la zone
    fusionnée décalée). _write_devis laisse déjà la ligne 3 vide par
    conception (le contenu commence à row=4) -- il suffit d'y écrire
    directement, sans jamais insérer de ligne ni toucher aux fusions."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["DEVIS QUANTITATIF"]
        ws.merge_cells("A3:G3")
        c = ws.cell(row=3, column=1, value=(
            "⚠ Recalcul automatique indisponible sur le serveur -- les colonnes Quantité/PU/Montant "
            "ci-dessous peuvent sembler vides tant que ce fichier n'a pas été OUVERT ET ENREGISTRÉ UNE "
            "FOIS dans Excel ou LibreOffice (recalcul automatique à l'ouverture). Les formules elles-mêmes "
            "sont correctes -- ouvre le fichier normalement pour voir les valeurs."
        ))
        c.font = Font(bold=True, size=10, color="884400")
        c.fill = PatternFill("solid", fgColor="FFF3CD")
        wb.save(str(xlsx_path))
    except Exception as e:
        print(f"[recalc xlsx] Échec de l'insertion du bandeau d'avertissement: {e}")

app = FastAPI()
app.mount("/outputs", StaticFiles(directory=str(OUTPUTS_DIR)), name="outputs")

# job_id -> {
#   "phase": "extraction"|"needs_input"|"reasoning"|"generation"|"done"|"error",
#   "progress": int, "stage": str, "done": bool, "error": str|None, "files": [...],
#   "missing_info": [...], "bilan": {...}, "boq": {...}, "answers": {...},
#   "project_name": str, "location": str,
# }
JOBS = {}


@app.get("/", response_class=HTMLResponse)
def index():
    return (BASE_DIR / "static" / "index.html").read_text(encoding="utf-8")


def _set(job_id, **kwargs):
    JOBS[job_id].update(kwargs)


def _finalize(job_id: str):
    """Devis déterministe + relecture Groq + génération Excel/PDF. Appelée
    une fois toutes les questions bloquantes résolues (ou s'il n'y en avait
    aucune)."""
    job = JOBS[job_id]
    try:
        _set(job_id, phase="reasoning", stage="Construction du devis chiffré...", progress=75)
        devis = build_devis(job["bilan"], KNOWLEDGE_BASE, job["project_name"], job["location"])

        _set(job_id, stage="Relecture IA (Groq)...", progress=85)
        devis["avertissements"] = review_devis(devis)

        _set(job_id, phase="generation", stage="Génération du fichier Excel...", progress=92)
        xlsx_path = OUTPUTS_DIR / f"{job_id}_devis.xlsx"
        generate_excel(job["bilan"], job["answers"], KNOWLEDGE_BASE, job["project_name"],
                        job["location"], devis, str(xlsx_path))
        recalc_ok = _recalculate_xlsx_with_libreoffice(xlsx_path)
        if not recalc_ok:
            _insert_recalc_warning_banner(xlsx_path)

        _set(job_id, stage="Génération du PDF bilan...", progress=97)
        pdf_path = OUTPUTS_DIR / f"{job_id}_bilan.pdf"
        generate_pdf(devis, str(pdf_path))

        json_path = OUTPUTS_DIR / f"{job_id}_devis.json"
        json_path.write_text(json.dumps({
            "devis": devis,
            "bilan": job["bilan"],
        }, ensure_ascii=False, indent=2), encoding="utf-8")

        _set(job_id, phase="done", progress=100, stage="Terminé.", done=True,
             files=[xlsx_path.name, pdf_path.name, json_path.name], devis=devis)
    except Exception as e:
        _set(job_id, phase="error", error=str(e), done=True)


def _run_job(job_id: str, pdf_path: str):
    def on_log(msg: str):
        JOBS[job_id]["stage"] = msg
        JOBS[job_id]["progress"] = min(60, JOBS[job_id]["progress"] + 2)

    try:
        _set(job_id, phase="extraction", stage="Analyse du document...", progress=2)
        result = run_pipeline(pdf_path, on_log=on_log)

        if not result.get("boq"):
            _set(job_id, phase="error", done=True,
                 error=result.get("avertissement", "Aucune donnée exploitable extraite."))
            return

        _set(job_id, bilan=result["bilan"], boq=result["boq"], progress=65,
             stage="Vérification des données manquantes...", pages_analysees=result["pages_analysees"])

        # v44 -- si un plan architectural a été fourni dès le départ, tente
        # la dérivation automatique (surface, longueur de réseau continu,
        # longueur de chaînage) MAINTENANT, avant même de savoir quelles
        # questions restent à poser -- évite de faire deviner un total à
        # l'utilisateur puis de lui demander une pièce jointe séparément en
        # cours de route pour la même info.
        archi_bytes = JOBS[job_id].get("archi_bytes")
        if archi_bytes:
            _set(job_id, stage="Lecture du plan architectural fourni...", progress=68)
            _try_derive_from_archi(job_id, archi_bytes, JOBS[job_id]["archi_filename"])

        questions = detect_missing_questions(JOBS[job_id]["bilan"])
        unanswered = [q for q in questions if q["key"] not in JOBS[job_id]["answers"]]
        if unanswered:
            _set(job_id, phase="needs_input", missing_info=unanswered,
                 stage="Informations manquantes -- complète les champs pour continuer.", progress=70)
            return

        _finalize(job_id)
    except Exception as e:
        _set(job_id, phase="error", error=str(e), done=True)


def _try_derive_from_archi(job_id: str, archi_bytes: bytes, archi_filename: str) -> None:
    """v44 -- tente les dérivations automatiques connues depuis le plan
    architectural fourni dès le départ. Best-effort: une dérivation qui
    échoue (Gemini renvoie null, ou erreur technique) n'empêche pas les
    autres d'être tentées, et laisse simplement la question correspondante
    apparaître normalement en aval, comme si aucun plan archi n'avait été
    fourni -- ne bloque jamais le job sur un échec de dérivation."""
    bilan = JOBS[job_id]["bilan"]
    candidats = []
    if bilan.get("surface_dallage", {}).get("donnee_indisponible", True) or \
       "approximation rectangle" in (bilan.get("surface_dallage", {}).get("source") or ""):
        candidats.append(("surface_batiment_totale_m2", "Surface totale du bâtiment (m²)"))
    if not bilan.get("longrines_par_section") and any(
        _normalise_type(t.get("type_designation")) not in EXCLUS_RESEAU_LONGRINES
        for t in bilan.get("longrines_reseau_continu", [])
    ):
        candidats.append(("longueur_reseau_longrine_totale", "Longueur développée totale du réseau de longrines (m)"))
    if bilan.get("chainage_types"):
        candidats.append(("longueur_chainage_totale", "Longueur développée totale du chaînage (m)"))

    for key, question_text in candidats:
        try:
            val = extract_value_from_attachment(archi_bytes, archi_filename, question_text, key=key)
            if val is not None:
                JOBS[job_id]["answers"][key] = float(val)
                print(f"[archi auto] {key} dérivé automatiquement depuis le plan architectural: {val}")
        except GeminiError as e:
            print(f"[archi auto] échec dérivation {key}: {e} -- la question sera posée normalement.")


@app.post("/api/run")
async def api_run(files: list[UploadFile] = File(...),
                   archi: UploadFile | None = File(None),
                   project_name: str = Form("Projet BTP"),
                   location: str = Form("")):
    """v44 -- accepte maintenant un second fichier optionnel 'archi' (plan
    architectural), demandé DÈS LE DÉPART plutôt que réclamé en cours de
    route via une pièce jointe reconstruite lors des questions. S'il est
    fourni, les valeurs qu'on sait en dériver (surface bâtiment, longueur
    de réseau continu, longueur de chaînage) sont tentées automatiquement
    juste après l'extraction, avant même de savoir si l'utilisateur devra
    répondre à quoi que ce soit -- voir _run_job."""
    if not files:
        return JSONResponse({"error": "Aucun fichier reçu."}, status_code=400)

    job_id = str(uuid.uuid4())
    pdf_file = files[0]
    pdf_path = UPLOADS_DIR / f"{job_id}_{pdf_file.filename}"
    with open(pdf_path, "wb") as f:
        shutil.copyfileobj(pdf_file.file, f)

    archi_bytes, archi_filename = None, None
    if archi is not None and getattr(archi, "filename", None):
        archi_bytes = await archi.read()
        archi_filename = archi.filename

    JOBS[job_id] = {
        "phase": "extraction", "progress": 0, "stage": "En file d'attente...",
        "done": False, "error": None, "files": [], "missing_info": [],
        "bilan": None, "boq": None, "answers": {}, "pages_analysees": [], "devis": None,
        "project_name": project_name, "location": location,
        "archi_bytes": archi_bytes, "archi_filename": archi_filename,
    }
    threading.Thread(target=_run_job, args=(job_id, str(pdf_path)), daemon=True).start()
    return {"job_id": job_id}


@app.post("/api/complete/{job_id}")
async def api_complete(job_id: str, request: Request):
    """Reçoit les réponses de l'utilisateur (texte et/ou pièces jointes,
    multipart/form-data) aux questions d'info manquante et relance le
    pipeline. Champs attendus par question de clé K : 'text__K' (optionnel)
    et 'file__K' (optionnel, fichier)."""
    job = JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Job introuvable."}, status_code=404)
    if job["phase"] != "needs_input":
        return JSONResponse({"error": "Ce job n'attend pas de complément d'info."}, status_code=400)

    form = await request.form()
    resolved = {}
    unresolved_keys = []

    for q in job["missing_info"]:
        key = q["key"]
        text_answer = form.get(f"text__{key}")
        upload = form.get(f"file__{key}")
        file_bytes, filename = None, None
        if upload is not None and hasattr(upload, "read"):
            file_bytes = await upload.read()
            filename = upload.filename

        value, source, debug_note = resolve_answer(key, q["question"], text_answer, file_bytes, filename,
                                                     kind=q.get("kind", "scalar"),
                                                     allow_attachment=q.get("allow_attachment", True))
        if value is None:
            if q.get("optional"):
                continue  # laissé vide volontairement -- pas bloquant, on garde le repli existant
            unresolved_keys.append(f"{key} ({debug_note})" if debug_note else key)
        else:
            resolved[key] = value

    if unresolved_keys:
        return JSONResponse(
            {"error": f"Impossible de résoudre: {', '.join(unresolved_keys)}."},
            status_code=400,
        )

    job["answers"].update(resolved)
    job["bilan"] = apply_answers_to_bilan(job["bilan"], job["answers"])
    job["done"] = False
    job["missing_info"] = []

    threading.Thread(target=_finalize, args=(job_id,), daemon=True).start()
    return {"ok": True}


@app.get("/api/status/{job_id}")
def api_status(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Job introuvable."}, status_code=404)
    # pages_analysees/devis peuvent être volumineux (des centaines de pages) --
    # jamais utiles au polling de statut (1x/seconde), seulement à /api/explore
    # à la demande une fois le job terminé.
    # v45 -- archi_bytes (contenu brut du PDF/image archi, en bytes) doit
    # AUSSI être exclu: FastAPI tente de sérialiser tout ce qui est renvoyé
    # ici en JSON, et son encodeur par défaut pour bytes fait un .decode()
    # UTF-8 -- un PDF binaire n'est jamais du texte UTF-8 valide, donc CETTE
    # route plantait en boucle (500) sur chaque poll de statut dès qu'un
    # plan archi avait été fourni au départ (v44), même après que la
    # dérivation automatique elle-même ait réussi.
    return {k: v for k, v in job.items()
            if k not in ("bilan", "boq", "pages_analysees", "devis", "archi_bytes", "archi_filename")}


@app.get("/api/explore/{job_id}")
def api_explore(job_id: str):
    """Tout ce dont l'explorateur en-app a besoin: devis chiffré ligne par
    ligne, éléments bruts extraits des plans, et pages analysées -- sans
    passer par le fichier Excel. Disponible une fois le job terminé."""
    job = JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Job introuvable."}, status_code=404)
    if job["phase"] != "done" or not job.get("devis"):
        return JSONResponse({"error": "Le job n'est pas encore terminé."}, status_code=400)

    bilan = job.get("bilan") or {}
    return {
        "project_name": job["project_name"], "location": job["location"],
        "devis": job["devis"],
        "pages_analysees": job.get("pages_analysees") or [],
        "elements": {
            "semelles": bilan.get("semelles", []),
            "radiers": bilan.get("radiers", []),
            "poteaux": bilan.get("poteaux", {}),
            "poteaux_coffrage_par_section": bilan.get("poteaux_coffrage_par_section", []),
            "raidisseurs_legende_par_section": bilan.get("raidisseurs_legende_par_section", []),
            "voiles_par_type": bilan.get("voiles_par_type", []),
            "longrines_par_section": bilan.get("longrines_par_section", []),
            "longrines_reseau_continu": bilan.get("longrines_reseau_continu", []),
            "escaliers": bilan.get("escaliers", []),
            "surface_dallage": bilan.get("surface_dallage", {}),
        },
        "answers": job.get("answers", {}),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
