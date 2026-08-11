"""
Génère les livrables finaux.

Le PDF (generate_pdf) est un bilan figé (snapshot) à partir du devis déjà
chiffré par devis_builder.py -- pratique pour un compte-rendu, pas pour être
modifié.

L'Excel (generate_excel) est construit différemment et EXPRÈS sans passer
par devis_builder : trois feuilles reliées par des FORMULES Excel, pour que
l'utilisateur puisse changer une épaisseur, une marge ou un prix unitaire
directement dans le fichier et voir tout se recalculer, sans repasser par ce
script Python.

  - "Paramètres"     : toutes les hypothèses confirmées (épaisseurs, marge de
                        fouille, hauteur de soubassement...) + les prix
                        unitaires, dans des cellules éditables.
  - "Bilan Éléments" : le détail brut par type d'ouvrage (semelles, poteaux,
                        longrines, voiles...) tel qu'extrait des plans, avec
                        les volumes/surfaces calculés par formule à partir
                        des paramètres.
  - "DEVIS QUANTITATIF" : les lignes officielles du devis, dont les cellules
                        quantité/PU/montant sont des formules pointant vers
                        les deux feuilles précédentes.
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from missing_info import _normalise_type as _normalise_type_reseau, EXCLUS_RESEAU_LONGRINES as EXCLUS_RESEAU_LONGRINES_XLSX

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from devis_template import SECTION_I_GENERALITES, SECTION_II_TERRASSEMENT, SECTION_III_BETON, SECTIONS_HORS_PERIMETRE, POSTE_KEY_TO_CODE

HEADER_FILL = "10b981"
SECTION_FILL = "e5e7eb"
WARN_FILL = "fff3cd"
PARAM_FILL = "e0f2ec"

# Valeurs de repli pour CHAQUE paramètre, pour que la feuille Paramètres les
# écrive TOUJOURS (même si la question n'a pas été posée -- ex: aucun voile
# détecté donc aucune question sur son épaisseur). Comme ça les formules de
# Bilan Éléments peuvent toujours référencer ces cellules, et si l'utilisateur
# ajoute plus tard une ligne manuellement (un voile que l'IA aurait raté par
# exemple), le calcul fonctionne immédiatement sans avoir à ressaisir un
# paramètre qui n'existerait pas encore.
DEFAULT_PARAMS = {
    "hauteur_soubassement_m": 0,
    "profondeur_ancrage_m": 0,
    "marge_fouille_pct": 15,
    "epaisseur_beton_proprete_cm": 10,
    "epaisseur_dallage_cm": 13,
    "epaisseur_voile_cm": 20,
    "volume_beton_banche_m3": 0,
}
# v25 -- bug corrigé: la feuille Excel attendait deux clés
# "largeur_semelle_filante_cm"/"hauteur_semelle_filante_cm" qui n'ont
# JAMAIS existé côté réponses utilisateur -- missing_info.py ne produit
# qu'une seule clé combinée "section_semelle_filante_cm" (ex: "40x20"),
# donc ce paramètre retombait TOUJOURS sur le repli codé en dur ci-dessous,
# quelle que soit la réponse réellement confirmée par l'utilisateur.
DEFAULT_SECTION_SEMELLE_FILANTE_CM = "40x20"


def _fmt_fcfa(n) -> str:
    if n is None:
        return "-"
    try:
        return f"{int(round(n)):,}".replace(",", " ")
    except (TypeError, ValueError):
        return str(n)


def _split_section(section: str):
    """'20x40' -> (20.0, 40.0) en cm, sinon (None, None)."""
    s = (section or "").strip().upper()
    if s.startswith("D"):
        try:
            d = float(s[1:])
            return d, d  # diamètre traité comme un carré équivalent grossier
        except ValueError:
            return None, None
    if "X" in s:
        try:
            a, b = s.split("X")
            return float(a), float(b)
        except ValueError:
            return None, None
    return None, None


# ------------------------------------------------------------------------
# Feuille "Paramètres"
# ------------------------------------------------------------------------

def _write_parametres(ws, answers: dict, kb: dict, header_fill, param_fill, border):
    ws["A1"] = "Paramètres de calcul (modifiables)"
    ws["A1"].font = Font(size=14, bold=True)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50

    row = 3
    ws.cell(row=row, column=1, value="Hypothèses confirmées par l'utilisateur").font = Font(bold=True, size=11)
    row += 1

    param_labels = {
        "hauteur_soubassement_m": ("Hauteur de soubassement (m)", "m"),
        "profondeur_ancrage_m": ("Profondeur d'ancrage des fondations (m)", "m"),
        "marge_fouille_pct": ("Marge de fouille (%)", "%"),
        "epaisseur_beton_proprete_cm": ("Épaisseur béton de propreté (cm)", "cm"),
        "epaisseur_dallage_cm": ("Épaisseur dallage (cm)", "cm"),
        "epaisseur_voile_cm": ("Épaisseur voile en soubassement (cm)", "cm"),
        "volume_beton_banche_m3": ("Volume béton banché/cyclopéen fondations filantes (m3, poste 3.2)", "m3"),
    }
    param_cells = {}
    for key, (label, unit) in param_labels.items():
        value = answers.get(key, DEFAULT_PARAMS[key])
        ws.cell(row=row, column=1, value=f"{label}").border = border
        c = ws.cell(row=row, column=2, value=value)
        c.border = border
        c.fill = param_fill
        c.font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"Modifie cette cellule pour recalculer le devis (unité: {unit}).").font = Font(italic=True, size=9, color="666666")
        param_cells[key] = f"'Paramètres'!${get_column_letter(2)}${row}"
        row += 2

    # v25 -- section semelle filante: une SEULE réponse utilisateur
    # ("section_semelle_filante_cm", ex: "40x20") mais deux paramètres
    # séparés nécessaires pour les formules Largeur/Hauteur ci-dessous.
    section_sf = answers.get("section_semelle_filante_cm") or DEFAULT_SECTION_SEMELLE_FILANTE_CM
    largeur_sf, hauteur_sf = _split_section(section_sf)
    if largeur_sf is None:
        largeur_sf, hauteur_sf = _split_section(DEFAULT_SECTION_SEMELLE_FILANTE_CM)
    for sub_key, sub_label, sub_val in [
        ("largeur_semelle_filante_cm", "Largeur semelle filante sous longrines (cm, poste 3.4)", largeur_sf),
        ("hauteur_semelle_filante_cm", "Hauteur semelle filante sous longrines (cm, poste 3.4)", hauteur_sf),
    ]:
        ws.cell(row=row, column=1, value=sub_label).border = border
        c = ws.cell(row=row, column=2, value=sub_val)
        c.border = border
        c.fill = param_fill
        c.font = Font(bold=True)
        ws.cell(row=row, column=3, value="Modifie cette cellule pour recalculer le devis (unité: cm).").font = Font(italic=True, size=9, color="666666")
        param_cells[sub_key] = f"'Paramètres'!${get_column_letter(2)}${row}"
        row += 2

    row += 1
    ws.cell(row=row, column=1, value="Prix unitaires (FCFA) — modifiables").font = Font(bold=True, size=11)
    row += 1
    headers = ["Code", "Désignation", "Unité", "PU (FCFA)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="ffffff")
        c.fill = header_fill
        c.border = border
    row += 1

    pu_cells = {}
    for code, info in kb.get("postes", {}).items():
        ws.cell(row=row, column=1, value=code).border = border
        ws.cell(row=row, column=2, value=info.get("designation", "")).border = border
        ws.cell(row=row, column=3, value=info.get("unite", "")).border = border
        c = ws.cell(row=row, column=4, value=info.get("prix_unitaire_fcfa"))
        c.border = border
        c.fill = param_fill
        pu_cells[code] = f"'Paramètres'!${get_column_letter(4)}${row}"
        row += 1

    return param_cells, pu_cells


# ------------------------------------------------------------------------
# Feuille "Bilan Éléments"
# ------------------------------------------------------------------------

def _write_bilan_elements(ws, bilan: dict, answers: dict, param_cells: dict, header_fill, border):
    for i, w in enumerate([26, 12, 12, 12, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = "Bilan des éléments (détail brut utilisé dans les calculs)"
    ws["A1"].font = Font(size=14, bold=True)
    row = 3
    total_cells = {}  # code -> "'Bilan Éléments'!$G$12"
    SHEET = "'Bilan Éléments'"

    def _table_header(row, cols):
        for i, h in enumerate(cols, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(bold=True, color="ffffff")
            c.fill = header_fill
            c.border = border
        return row + 1

    # ---- Semelles isolées (3.1 béton propreté + 3.3 béton armé) ----
    ws.cell(row=row, column=1, value="Semelles isolées").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Désignation", "a (m)", "b (m)", "h (cm)", "Nombre",
                               "Vol. béton armé (m3)", "Vol. béton propreté (m3)"])
    first_row = row
    for s in bilan.get("semelles", []):
        ws.cell(row=row, column=1, value=s.get("designation", "")).border = border
        ws.cell(row=row, column=2, value=s["a_m"]).border = border
        ws.cell(row=row, column=3, value=s["b_m"]).border = border
        ws.cell(row=row, column=4, value=s["h_cm"]).border = border
        ws.cell(row=row, column=5, value=s["nombre"]).border = border
        r = row
        ws.cell(row=row, column=6, value=f"=B{r}*C{r}*(D{r}/100)*E{r}").border = border
        ws.cell(row=row, column=7, value=f"=B{r}*C{r}*({param_cells['epaisseur_beton_proprete_cm']}/100)*E{r}").border = border
        row += 1
    if row == first_row:  # aucune semelle détectée -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter une semelle ici)").font = Font(italic=True, color="999999")
        for col, val in ((2, 0), (3, 0), (4, 0), (5, 0)):
            ws.cell(row=row, column=col, value=val).border = border
        r = row
        ws.cell(row=row, column=6, value=f"=B{r}*C{r}*(D{r}/100)*E{r}").border = border
        ws.cell(row=row, column=7, value=f"=B{r}*C{r}*({param_cells['epaisseur_beton_proprete_cm']}/100)*E{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"=SUM(G{first_row}:G{last_row})").font = Font(bold=True)
    total_cells["3.3"] = f"{SHEET}!$F${row}"
    total_cells["3.1"] = f"{SHEET}!$G${row}"
    row += 3

    # ---- Semelles filantes (3.4) : longueur développée = longueur totale
    # des longrines (référence directe à la table Longrines ci-dessous),
    # section confirmée par l'utilisateur en Paramètres. Note: la table
    # Longrines est écrite juste après pour connaître longrines_longueur_cell
    # -- on réserve donc la ligne ici et on la complète plus bas. ----
    row_semelles_filantes = row
    ws.cell(row=row, column=1, value="Semelles filantes (sous longrines)").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Longueur développée (m)", "Largeur section (m, param.)",
                               "Hauteur section (m, param.)", "Volume béton (m3)", "", "", ""])
    row_sf_data = row
    row += 1
    row += 2

    # ---- Poteaux (3.5 potelets, dépend de la hauteur soubassement) ----
    ws.cell(row=row, column=1, value="Poteaux (potelets en soubassement)").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Section", "Nombre", "Aire section (m2)", "Hauteur (m, param.)", "Volume béton (m3)", "", ""])
    first_row = row
    hc = param_cells["hauteur_soubassement_m"]
    # v25 -- priorité de source : (1) comptage individuel par section, sinon
    # (2) répartition EXACTE lue en légende (colonne Quantité, voir
    # pipeline.py: total_legende_par_section), sinon (3) total global +
    # section représentative confirmée par l'utilisateur. Avant ce correctif,
    # cette table ne lisait QUE la source (1) -- sur une grille dense sans
    # comptage individuel fiable, elle restait vide même quand (2) ou (3)
    # avaient bien été calculés côté bilan/réponses utilisateur.
    _poteaux_rows = (
        bilan.get("poteaux", {}).get("par_section")
        or bilan.get("poteaux", {}).get("total_legende_par_section")
        or []
    )
    if not _poteaux_rows:
        _total_global = bilan.get("poteaux", {}).get("total_legende_global")
        _section_confirmee = answers.get("section_poteaux_total_global_cm")
        if _total_global and _section_confirmee:
            _poteaux_rows = [{"section": _section_confirmee, "nombre_total": _total_global}]
    for item in _poteaux_rows:
        a_cm, b_cm = _split_section(item["section"])
        aire = round((a_cm / 100) * (b_cm / 100), 4) if a_cm and b_cm else None
        ws.cell(row=row, column=1, value=item["section"]).border = border
        ws.cell(row=row, column=2, value=item["nombre_total"]).border = border
        ws.cell(row=row, column=3, value=aire).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={hc}").border = border
        ws.cell(row=row, column=5, value=f"=C{r}*B{r}*D{r}").border = border
        row += 1
    if row == first_row:  # aucun poteau détecté -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter un poteau ici, ex: 20x20)").font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=0).border = border
        ws.cell(row=row, column=3, value=0).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={hc}").border = border
        ws.cell(row=row, column=5, value=f"=C{r}*B{r}*D{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{first_row}:E{last_row})").font = Font(bold=True)
    total_cells["3.5"] = f"{SHEET}!$E${row}"
    row += 3

    # ---- v38 -- Raidisseurs (3.5bis, poste ajouté hors canevas standard --
    # voir pipeline.py/devis_template.py). Même hauteur de soubassement que
    # les potelets ci-dessus. Table absente jusqu'ici -- total_cells n'avait
    # donc jamais de clé "3.5bis", même quand le poste existait côté bilan. ----
    _raidisseurs_rows = bilan.get("raidisseurs_legende_par_section") or []
    if _raidisseurs_rows:
        ws.cell(row=row, column=1, value="Raidisseurs (niveau soubassement -- poste ajouté 3.5bis)").font = Font(bold=True, size=12)
        row += 1
        row = _table_header(row, ["Désignation", "Section", "Nombre", "Aire section (m2)", "Hauteur (m, param.)", "Volume béton (m3)", ""])
        first_row = row
        for item in _raidisseurs_rows:
            a_cm, b_cm = _split_section(item["section"])
            aire = round((a_cm / 100) * (b_cm / 100), 4) if a_cm and b_cm else None
            ws.cell(row=row, column=1, value=item["designation"]).border = border
            ws.cell(row=row, column=2, value=item["section"]).border = border
            ws.cell(row=row, column=3, value=item["nombre_total"]).border = border
            ws.cell(row=row, column=4, value=aire).border = border
            r = row
            ws.cell(row=row, column=5, value=f"={hc}").border = border
            ws.cell(row=row, column=6, value=f"=D{r}*C{r}*E{r}").border = border
            row += 1
        last_row = row - 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
        total_cells["3.5bis"] = f"{SHEET}!$F${row}"
        row += 3

    # ---- Voiles (3.6, dépend de hauteur + épaisseur voile) ----
    ws.cell(row=row, column=1, value="Voiles en soubassement").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Désignation", "Nombre", "Longueur totale (m)", "Épaisseur (m, param.)", "Hauteur (m, param.)", "Volume béton (m3)", ""])
    first_row = row
    ec = param_cells["epaisseur_voile_cm"]
    for v in bilan.get("voiles_par_type", []):
        ws.cell(row=row, column=1, value=v.get("designation", "")).border = border
        ws.cell(row=row, column=2, value=v.get("nombre")).border = border
        ws.cell(row=row, column=3, value=v["longueur_totale_m"]).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={ec}/100").border = border
        ws.cell(row=row, column=5, value=f"={hc}").border = border
        ws.cell(row=row, column=6, value=f"=C{r}*D{r}*E{r}").border = border
        row += 1
    if row == first_row:  # aucun voile détecté -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter un voile ici)").font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=0).border = border
        ws.cell(row=row, column=3, value=0).border = border
        r = row
        ws.cell(row=row, column=4, value=f"={ec}/100").border = border
        ws.cell(row=row, column=5, value=f"={hc}").border = border
        ws.cell(row=row, column=6, value=f"=C{r}*D{r}*E{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{first_row}:B{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(C{first_row}:C{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
    total_cells["3.6"] = f"{SHEET}!$F${row}"
    voiles_longueur_totale_cell = f"{SHEET}!$C${row}"
    row += 3

    # ---- Longrines (3.7, + largeur cumulée pour fouilles 2.4) ----
    ws.cell(row=row, column=1, value="Longrines").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Section", "Tronçons", "Longueur totale (m)", "Largeur (m)", "Volume béton (m3)", "Largeur x Longueur (m2)", ""])
    first_row = row
    # ---- v34 -- repli réseau continu : si aucun tronçon individuellement
    # désigné (bilan["longrines_par_section"] vide) mais un réseau continu
    # détecté en légende ET une longueur totale confirmée par l'utilisateur
    # (longueur_reseau_longrine_totale -- v32, une seule longueur pour tout
    # le réseau, appliquée à la section la plus grosse détectée), construit
    # une ligne unique à partir de cette confirmation. Ce bloc pointait
    # encore vers l'ancienne clé de réponse (longueur_totale_reseau_
    # longrines_m, abandonnée depuis v29/v32) -- jamais mise à jour lors des
    # refontes de missing_info.py, donc cette table restait vide même
    # quand missing_info.py avait bien calculé le volume correspondant.
    _longrines_rows = bilan.get("longrines_par_section") or []
    if not _longrines_rows:
        _longueur_confirmee = answers.get("longueur_reseau_longrine_totale")
        _types_valides = [
            t for t in (bilan.get("longrines_reseau_continu") or [])
            if _normalise_type_reseau(t.get("type_designation")) not in EXCLUS_RESEAU_LONGRINES_XLSX
        ]
        if _longueur_confirmee and _types_valides:
            _section_max_item = max(
                (t for t in _types_valides if _split_section(t["section"])[0] is not None),
                key=lambda t: (lambda a, b: a * b)(*_split_section(t["section"])),
                default=None,
            )
            if _section_max_item:
                _longrines_rows = [{
                    "section": _section_max_item["section"],
                    "nombre_troncons": None,
                    "longueur_totale_m": _longueur_confirmee,
                }]
    for item in _longrines_rows:
        a_cm, b_cm = _split_section(item["section"])
        largeur_m = round(min(a_cm, b_cm) / 100, 4) if a_cm and b_cm else None
        aire_section = round((a_cm / 100) * (b_cm / 100), 4) if a_cm and b_cm else None
        ws.cell(row=row, column=1, value=item["section"]).border = border
        ws.cell(row=row, column=2, value=item.get("nombre_troncons")).border = border
        ws.cell(row=row, column=3, value=item["longueur_totale_m"]).border = border
        ws.cell(row=row, column=4, value=largeur_m).border = border
        r = row
        ws.cell(row=row, column=5, value=(f"={aire_section}*C{r}" if aire_section is not None else None)).border = border
        ws.cell(row=row, column=6, value=f"=D{r}*C{r}").border = border
        row += 1
    if row == first_row:  # aucune longrine détectée -- une ligne modèle éditable
        ws.cell(row=row, column=1, value="(ajouter une longrine ici, ex: 20x30)").font = Font(italic=True, color="999999")
        ws.cell(row=row, column=2, value=0).border = border
        ws.cell(row=row, column=3, value=0).border = border
        ws.cell(row=row, column=4, value=0).border = border
        r = row
        ws.cell(row=row, column=5, value=f"=D{r}*C{r}").border = border  # largeur * longueur (approx, à corriger si section réelle connue)
        ws.cell(row=row, column=6, value=f"=D{r}*C{r}").border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(C{first_row}:C{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{first_row}:E{last_row})").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
    total_cells["3.7"] = f"{SHEET}!$E${row}"
    longrines_longueur_totale_cell = f"{SHEET}!$C${row}"
    longrines_largeur_longueur_cell = f"{SHEET}!$F${row}"
    row += 3

    # ---- Complète la ligne "Semelles filantes" réservée plus haut, maintenant
    # que la longueur développée totale des longrines est connue ----
    ws.cell(row=row_sf_data, column=1, value=f"={longrines_longueur_totale_cell}").border = border
    ws.cell(row=row_sf_data, column=2, value=f"={param_cells['largeur_semelle_filante_cm']}/100").border = border
    ws.cell(row=row_sf_data, column=3, value=f"={param_cells['hauteur_semelle_filante_cm']}/100").border = border
    ws.cell(row=row_sf_data, column=4,
            value=f"=A{row_sf_data}*B{row_sf_data}*C{row_sf_data}").border = border
    total_cells["3.4"] = f"{SHEET}!$D${row_sf_data}"

    # ---- Fouilles (2.3 puits, 2.4 rigoles) : dépendent de profondeur + marge ----
    ws.cell(row=row, column=1, value="Fouilles").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Type", "Emprise (m2)", "Profondeur (m, param.)", "Marge (%, param.)", "Volume (m3)", "", ""])
    surface_semelles = sum(s["a_m"] * s["b_m"] * s["nombre"] for s in bilan.get("semelles", []))
    if "profondeur_ancrage_m" in param_cells and "marge_fouille_pct" in param_cells:
        pc = param_cells["profondeur_ancrage_m"]
        mc = param_cells["marge_fouille_pct"]
        r = row
        ws.cell(row=row, column=1, value="Puits (semelles isolées)").border = border
        ws.cell(row=row, column=2, value=round(surface_semelles, 3)).border = border
        ws.cell(row=row, column=3, value=f"={pc}").border = border
        ws.cell(row=row, column=4, value=f"={mc}").border = border
        ws.cell(row=row, column=5, value=f"=B{r}*C{r}*(1+D{r}/100)").border = border
        total_cells["2.3"] = f"{SHEET}!$E${r}"
        row += 1
        r = row
        ws.cell(row=row, column=1, value="Rigoles (longrines)").border = border
        ws.cell(row=row, column=2, value=f"={longrines_largeur_longueur_cell}").border = border
        ws.cell(row=row, column=3, value=f"={pc}").border = border
        ws.cell(row=row, column=4, value=f"={mc}").border = border
        ws.cell(row=row, column=5, value=f"=B{r}*C{r}*(1+D{r}/100)").border = border
        total_cells["2.4"] = f"{SHEET}!$E${r}"
        row += 1
    else:
        total_cells["2.3"] = None
        total_cells["2.4"] = None
    row += 2

    # ---- Dallage (3.8) : surface brute - déductions, dépend épaisseur dallage + épaisseur voile ----
    sd = bilan.get("surface_dallage", {})
    ws.cell(row=row, column=1, value="Dallage").font = Font(bold=True, size=12)
    row += 1
    if not sd.get("donnee_indisponible", True):
        surface_brute = sd.get("surface_brute_m2", 0)
        deductions = sd.get("deductions_m2", {})
        ddp = deductions.get("poteaux", 0)
        ddl = deductions.get("longrines", 0)
        ddv_brute = deductions.get("voiles", 0)  # calculée avec épaisseur voile 20cm d'origine

        ws.cell(row=row, column=1, value="Surface brute (m2)").border = border
        ws.cell(row=row, column=2, value=surface_brute).border = border
        r_brute = row
        row += 1
        ws.cell(row=row, column=1, value="Déduction emprise poteaux (m2)").border = border
        ws.cell(row=row, column=2, value=ddp).border = border
        r_dp = row
        row += 1
        ws.cell(row=row, column=1, value="Déduction emprise longrines (m2)").border = border
        ws.cell(row=row, column=2, value=ddl).border = border
        r_dl = row
        row += 1
        ws.cell(row=row, column=1, value="Déduction emprise voiles (m2, mise à l'échelle épaisseur param.)").border = border
        if "epaisseur_voile_cm" in param_cells and ddv_brute:
            ec = param_cells["epaisseur_voile_cm"]
            ws.cell(row=row, column=2, value=f"={ddv_brute}*({ec}/20)").border = border
        else:
            ws.cell(row=row, column=2, value=ddv_brute).border = border
        r_dv = row
        row += 1
        ws.cell(row=row, column=1, value="Surface nette dallage (m2)").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"=B{r_brute}-B{r_dp}-B{r_dl}-B{r_dv}").font = Font(bold=True)
        r_nette = row
        row += 1
        ws.cell(row=row, column=1, value="Volume béton dallage (m3)").font = Font(bold=True)
        if "epaisseur_dallage_cm" in param_cells:
            edc = param_cells["epaisseur_dallage_cm"]
            ws.cell(row=row, column=2, value=f"=B{r_nette}*({edc}/100)").font = Font(bold=True)
            total_cells["3.8"] = f"{SHEET}!$B${row}"
        else:
            ws.cell(row=row, column=2, value=None).font = Font(bold=True)
            total_cells["3.8"] = None
        row += 1
    else:
        ws.cell(row=row, column=1, value=f"Indisponible: {sd.get('raison', 'surface dallage non calculable')}").font = Font(italic=True, color="884400")
        total_cells["3.8"] = None
        row += 1

    # ---- Escaliers (3.9 marches, 3.10 bèche) : pas de paramètre externe requis ----
    ws.cell(row=row, column=1, value="Escaliers").font = Font(bold=True, size=12)
    row += 1
    row = _table_header(row, ["Désignation", "Nb marches", "Giron (m)", "Hauteur marche (m)",
                               "Largeur volée (m)", "Vol. marches (m3)", "Vol. bèche (m3)"])
    first_row = row
    escaliers = bilan.get("escaliers", [])
    for e in escaliers:
        ws.cell(row=row, column=1, value=e.get("designation", "")).border = border
        ws.cell(row=row, column=2, value=e.get("nombre_marches")).border = border
        giron_m = round(e["giron_cm"] / 100, 4) if e.get("giron_cm") else None
        hm_m = round(e["hauteur_marche_cm"] / 100, 4) if e.get("hauteur_marche_cm") else None
        ws.cell(row=row, column=3, value=giron_m).border = border
        ws.cell(row=row, column=4, value=hm_m).border = border
        ws.cell(row=row, column=5, value=e.get("largeur_volee_m")).border = border
        r = row
        if e.get("nombre_marches") and giron_m and hm_m and e.get("largeur_volee_m"):
            ws.cell(row=row, column=6, value=f"=B{r}*(C{r}*D{r}/2)*E{r}").border = border
        else:
            ws.cell(row=row, column=6, value=None).border = border
        beche = e.get("beche") or {}
        if beche.get("longueur_m") and beche.get("largeur_cm") and beche.get("hauteur_cm"):
            ws.cell(row=row, column=7, value=round(
                (beche["largeur_cm"] / 100) * (beche["hauteur_cm"] / 100) * beche["longueur_m"], 4
            )).border = border
        else:
            ws.cell(row=row, column=7, value=None).border = border
        row += 1
    last_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    if last_row >= first_row:
        ws.cell(row=row, column=6, value=f"=SUM(F{first_row}:F{last_row})").font = Font(bold=True)
        ws.cell(row=row, column=7, value=f"=SUM(G{first_row}:G{last_row})").font = Font(bold=True)
    else:
        ws.cell(row=row, column=6, value=0).font = Font(bold=True)
        ws.cell(row=row, column=7, value=0).font = Font(bold=True)
    total_cells["3.9"] = f"{SHEET}!$F${row}" if any(
        e.get("nombre_marches") and e.get("giron_cm") and e.get("hauteur_marche_cm") and e.get("largeur_volee_m")
        for e in escaliers
    ) else None
    total_cells["3.10"] = f"{SHEET}!$G${row}" if any(
        (e.get("beche") or {}).get("longueur_m") for e in escaliers
    ) else None
    row += 3

    return total_cells


# ------------------------------------------------------------------------
# Feuille "DEVIS QUANTITATIF"
# ------------------------------------------------------------------------

def _write_devis(ws, devis: dict, header_fill, section_fill, warn_fill, border):
    """v39 -- réécrit pour transcrire directement les lignes déjà chiffrées
    de `devis` (le même dict que le JSON/PDF, construit une seule fois par
    devis_builder.build_devis) au lieu de reconstruire ses propres
    Quantité/PU/Montant via des formules cross-feuilles (=total_cells[...]).

    Avant ce correctif, CETTE feuille -- celle que le client ouvre en
    premier -- ne contenait AUCUNE valeur en dur: dès que le recalcul
    LibreOffice échouait (environnement de déploiement sans LibreOffice,
    ou tout autre accroc), TOUT redevenait vide d'un coup (quantités, prix,
    montants), même quand chaque calcul sous-jacent était juste. Les
    valeurs sont maintenant écrites telles quelles, indépendamment de tout
    recalcul -- ce fichier ne peut plus dépendre de savoir si LibreOffice
    est installé sur le serveur ou non."""
    col_widths = [8, 46, 8, 12, 14, 16, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    ws["A1"] = f"DEVIS QUANTITATIF ET ESTIMATIF — {devis['projet']}"
    ws["A1"].font = Font(size=15, bold=True)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{devis['localisation']} — {date.today().isoformat()} — Infrastructure uniquement (voir notes)"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    row = 4
    headers = ["Code", "Désignation des ouvrages", "Unité", "Quantité", "PU (FCFA)", "Montant (FCFA)", "Note"]

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=f"I. {SECTION_I_GENERALITES['titre']}")
    c.font = Font(bold=True, size=12)
    c.fill = section_fill
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=f"⚠ {SECTION_I_GENERALITES['note']}")
    c.font = Font(italic=True, size=9, color="884400")
    c.fill = warn_fill
    row += 2

    montant_total_general = 0
    for section in devis["sections"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f"{section['numero']}. {section['titre']}")
        c.font = Font(bold=True, size=12)
        c.fill = section_fill
        row += 1
        if section.get("note"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            c = ws.cell(row=row, column=1, value=f"⚠ {section['note']}")
            c.font = Font(italic=True, size=9, color="884400")
            c.fill = warn_fill
            row += 1

        def _write_header():
            nonlocal row
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=i, value=h)
                c.font = Font(bold=True, color="ffffff")
                c.fill = header_fill
                c.border = border
                c.alignment = Alignment(horizontal="center")
            row += 1

        def _write_lignes(lignes):
            nonlocal row
            first_data_row = row
            for ligne in lignes:
                indispo = ligne["source"] == "indisponible"

                ws.cell(row=row, column=1, value=ligne["code"]).border = border
                ws.cell(row=row, column=2, value=ligne["designation"]).border = border
                ws.cell(row=row, column=3, value=ligne["unite"]).border = border

                if indispo:
                    cq = ws.cell(row=row, column=4, value="À COMPLÉTER")
                    cpu = ws.cell(row=row, column=5, value=ligne["prix_unitaire_fcfa"] if ligne["prix_unitaire_fcfa"] else "-")
                    cm_ = ws.cell(row=row, column=6, value="-")
                    note_cell = ws.cell(row=row, column=7, value=ligne["note"] or "Non calculable (donnée hors périmètre ou question non posée).")
                    for cc in (cq, cpu, cm_, note_cell):
                        cc.fill = warn_fill
                else:
                    cq = ws.cell(row=row, column=4, value=ligne["quantite"])
                    cpu = ws.cell(row=row, column=5, value=ligne["prix_unitaire_fcfa"])
                    cm_ = ws.cell(row=row, column=6, value=ligne["montant_fcfa"])
                    note_cell = ws.cell(row=row, column=7, value=ligne["note"] or "")
                for cc in (cq, cpu, cm_, note_cell):
                    cc.border = border
                    if cc.column in (4, 5, 6):
                        cc.alignment = Alignment(horizontal="right")
                row += 1
            return first_data_row, row - 1

        # v47 -- si la section porte une ventilation par sous-section
        # (ex: III -> Infrastructures / Superstructures, comme sur le
        # canevas de référence), affiche chaque sous-groupe avec son propre
        # titre + en-têtes, au lieu d'un seul bloc fourre-tout.
        if section.get("sous_sections"):
            for sous in section["sous_sections"]:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                c = ws.cell(row=row, column=1, value=sous["titre"])
                c.font = Font(bold=True, size=11, italic=True)
                row += 1
                _write_header()
                _write_lignes(sous["lignes"])
                row += 1  # ligne blanche entre sous-sections
        else:
            _write_header()
            _write_lignes(section["lignes"])

        sous_total = sum(l["montant_fcfa"] for l in section["lignes"] if l["montant_fcfa"])
        montant_total_general += sous_total
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value="Sous-total section")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        c2 = ws.cell(row=row, column=6, value=sous_total)
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="right")
        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value="TOTAL INFRASTRUCTURE (postes chiffrés, HTVA)")
    c.font = Font(bold=True, size=13)
    c.fill = section_fill
    c.alignment = Alignment(horizontal="right")
    c2 = ws.cell(row=row, column=6, value=montant_total_general)
    c2.font = Font(bold=True, size=13)
    c2.fill = section_fill
    c2.alignment = Alignment(horizontal="right")
    row += 3

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="Autres lots du devis (hors périmètre de cette extraction automatique) :").font = Font(bold=True, size=11)
    row += 1
    for s in SECTIONS_HORS_PERIMETRE:
        ws.cell(row=row, column=1, value=f"{s['numero']}. {s['titre']} — à compléter manuellement").font = Font(italic=True, size=9, color="666666")
        row += 1
    row += 1

    avertissements = devis.get("avertissements") or []
    if avertissements:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value="Observations (relecture IA) :").font = Font(bold=True, size=11)
        row += 1
        for a in avertissements:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(row=row, column=1, value=f"- {a}").font = Font(italic=True, size=9, color="444444")
            row += 1


def generate_excel(bilan: dict, answers: dict, kb: dict, project_name: str, location: str,
                    devis: dict, out_path: str):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    section_fill = PatternFill("solid", fgColor=SECTION_FILL)
    warn_fill = PatternFill("solid", fgColor=WARN_FILL)
    param_fill = PatternFill("solid", fgColor=PARAM_FILL)
    thin = Side(style="thin", color="cccccc")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_param = wb.active
    ws_param.title = "Paramètres"
    param_cells, pu_cells = _write_parametres(ws_param, answers, kb, header_fill, param_fill, border)

    ws_elem = wb.create_sheet("Bilan Éléments")
    _write_bilan_elements(ws_elem, bilan, answers, param_cells, header_fill, border)

    ws_devis = wb.create_sheet("DEVIS QUANTITATIF")
    _write_devis(ws_devis, devis, header_fill, section_fill, warn_fill, border)

    wb.save(out_path)
    return out_path


# ------------------------------------------------------------------------
# PDF bilan (snapshot figé, à partir du devis déjà chiffré par devis_builder.py)
# ------------------------------------------------------------------------

def generate_pdf(devis: dict, out_path: str):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleFR", parent=styles["Title"], fontSize=16)
    sub_style = ParagraphStyle("SubFR", parent=styles["Normal"], textColor=colors.grey, fontSize=9)
    section_style = ParagraphStyle("SectionFR", parent=styles["Heading2"], fontSize=12, spaceBefore=14)
    note_style = ParagraphStyle("NoteFR", parent=styles["Normal"], textColor=colors.HexColor("#884400"), fontSize=8, spaceAfter=6)
    warn_style = ParagraphStyle("WarnFR", parent=styles["Normal"], textColor=colors.HexColor("#884400"), fontSize=8)

    doc = SimpleDocTemplate(out_path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
        Paragraph(f"Bilan — Devis Quantitatif Infrastructure — {devis['projet']}", title_style),
        Paragraph(f"{devis['localisation']} — {date.today().isoformat()}", sub_style),
        Spacer(1, 10),
    ]

    for section in devis["sections"]:
        story.append(Paragraph(f"{section['numero']}. {section['titre']}", section_style))
        if section.get("note"):
            story.append(Paragraph(f"⚠ {section['note']}", note_style))

        data = [["Code", "Désignation", "Unité", "Qté", "PU (FCFA)", "Montant (FCFA)"]]
        section_total = 0
        for ligne in section["lignes"]:
            indispo = ligne["source"] == "indisponible"
            data.append([
                ligne["code"], ligne["designation"][:60] + ("…" if len(ligne["designation"]) > 60 else ""),
                ligne["unite"],
                "À compléter" if indispo else str(ligne["quantite"]),
                _fmt_fcfa(ligne["prix_unitaire_fcfa"]) if ligne["prix_unitaire_fcfa"] else "-",
                "-" if indispo else _fmt_fcfa(ligne["montant_fcfa"]),
            ])
            if ligne["montant_fcfa"]:
                section_total += ligne["montant_fcfa"]
        data.append(["", "", "", "", "Sous-total", _fmt_fcfa(section_total)])

        table = Table(data, colWidths=[1.4 * cm, 6.8 * cm, 1.3 * cm, 2.0 * cm, 2.6 * cm, 3.0 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -2), 0.4, colors.HexColor("#cccccc")),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        story.append(Spacer(1, 6))

    story.append(Spacer(1, 10))
    total_style = ParagraphStyle("TotalFR", parent=styles["Heading2"], alignment=2)
    story.append(Paragraph(
        f"TOTAL INFRASTRUCTURE (HTVA) : {_fmt_fcfa(devis['total_infrastructure_fcfa'])} FCFA", total_style))

    story.append(Spacer(1, 14))
    story.append(Paragraph("Autres lots (hors périmètre de cette extraction, à compléter manuellement) :", section_style))
    for s in SECTIONS_HORS_PERIMETRE:
        story.append(Paragraph(f"- {s['numero']}. {s['titre']}", warn_style))

    if devis.get("postes_a_completer_manuellement"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("Points à compléter / confirmer :", section_style))
        for item in devis["postes_a_completer_manuellement"]:
            story.append(Paragraph(f"- {item['poste']} : {item['raison']}", warn_style))

    if devis.get("avertissements"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("Observations (relecture IA) :", section_style))
        for a in devis["avertissements"]:
            story.append(Paragraph(f"- {a}", warn_style))

    doc.build(story)
    return out_path
